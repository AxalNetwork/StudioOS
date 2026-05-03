"""Task #26 — Financial Model Builder.

Stores per-project 3-statement-style assumptions in `financial_models` and
computes runway / burn / breakeven plus a ±20% sensitivity grid on the
top-3 drivers. On save, derives capital-category slider values that the
v2 scoring engine (`services/scoring.py`) can consume.

Endpoints (all under `/api/financials`):
  GET    /{project_id}              — fetch model (or defaults if none yet)
  PUT    /{project_id}              — upsert model + recompute everything
  POST   /{project_id}/recompute    — recompute from stored assumptions
  GET    /{project_id}/export.xlsx  — download XLSX
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field as PydField
from sqlmodel import Session, select

from backend.app.api.deps import can_access_founder_resource, is_privileged
from backend.app.api.routes.auth import get_current_user
from backend.app.database import get_session
from backend.app.models.entities import (
    ActivityLog,
    FinancialModel,
    Project,
    User,
    UserRole,
)

router = APIRouter(prefix="/financials", tags=["Financial Model"])


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------
class Assumptions(BaseModel):
    """Drivers a founder edits. Defaults give a sane starter model."""
    starting_cash: float = PydField(default=250_000, ge=0)
    price_per_unit: float = PydField(default=99, ge=0)
    units_month_0: float = PydField(default=50, ge=0)
    monthly_growth_pct: float = PydField(default=12, ge=-100, le=200)
    cac: float = PydField(default=80, ge=0)
    monthly_churn_pct: float = PydField(default=4, ge=0, le=100)
    salaries_monthly: float = PydField(default=18_000, ge=0)
    opex_monthly: float = PydField(default=4_500, ge=0)
    gross_margin_pct: float = PydField(default=70, ge=0, le=100)
    horizon_months: int = PydField(default=24, ge=3, le=60)


# ---------------------------------------------------------------------------
# Computation core
# ---------------------------------------------------------------------------
def _project_financials(a: Assumptions) -> dict:
    """Run the per-month projection and aggregate runway / burn / breakeven."""
    months = []
    cash = float(a.starting_cash)
    units = float(a.units_month_0)
    growth = a.monthly_growth_pct / 100.0
    churn = a.monthly_churn_pct / 100.0
    gm = a.gross_margin_pct / 100.0

    breakeven_month: Optional[int] = None
    runway_months: Optional[int] = None
    burns: list[float] = []
    revenues: list[float] = []

    for m in range(1, a.horizon_months + 1):
        # New units acquired this month, then existing base churns.
        new_units = units * max(growth, 0)
        units = max(units * (1 - churn) + new_units, 0)
        revenue = units * a.price_per_unit
        gross_profit = revenue * gm
        marketing = new_units * a.cac
        fixed = a.salaries_monthly + a.opex_monthly
        total_cost = marketing + fixed
        net = gross_profit - total_cost
        cash += net
        burns.append(round(total_cost - gross_profit, 2))  # positive = burning
        revenues.append(round(revenue, 2))

        if breakeven_month is None and net >= 0:
            breakeven_month = m
        if runway_months is None and cash <= 0:
            runway_months = m

        months.append({
            "month": m,
            "units": round(units, 2),
            "revenue": round(revenue, 2),
            "gross_profit": round(gross_profit, 2),
            "marketing": round(marketing, 2),
            "fixed_cost": round(fixed, 2),
            "net": round(net, 2),
            "cash": round(cash, 2),
        })

    avg_burn = round(sum(b for b in burns if b > 0) / max(len([b for b in burns if b > 0]), 1), 2)
    last_cash = months[-1]["cash"] if months else float(a.starting_cash)

    # If cash never went negative within horizon, runway = horizon+ (use sentinel)
    if runway_months is None:
        if avg_burn > 0:
            runway_months_est = round(a.starting_cash / avg_burn, 1)
        else:
            runway_months_est = float(a.horizon_months)
        runway_value: float = max(runway_months_est, float(a.horizon_months))
        runway_capped = False
    else:
        runway_value = float(runway_months)
        runway_capped = True

    # LTV / CAC quick read.
    if churn > 0:
        ltv = (a.price_per_unit * gm) / churn
    else:
        ltv = a.price_per_unit * gm * a.horizon_months
    ltv_cac = round(ltv / a.cac, 2) if a.cac > 0 else None

    return {
        "months": months,
        "runway_months": round(runway_value, 1),
        "runway_capped": runway_capped,
        "avg_monthly_burn": avg_burn,
        "breakeven_month": breakeven_month,
        "ending_cash": round(last_cash, 2),
        "total_revenue_horizon": round(sum(revenues), 2),
        "ltv": round(ltv, 2),
        "ltv_cac_ratio": ltv_cac,
    }


def _sensitivity(a: Assumptions) -> dict:
    """±20% on the top-3 drivers: price, units_month_0, cac.
    Each row reports runway_months and breakeven_month.
    """
    drivers = [
        ("price_per_unit", "Price per unit"),
        ("units_month_0", "Starting units"),
        ("cac", "CAC"),
    ]
    deltas = [-0.20, -0.10, 0.0, 0.10, 0.20]
    rows = []
    for key, label in drivers:
        cells = []
        for d in deltas:
            payload = a.model_dump()
            payload[key] = max(0.0, payload[key] * (1 + d))
            r = _project_financials(Assumptions(**payload))
            cells.append({
                "delta_pct": int(d * 100),
                "runway_months": r["runway_months"],
                "breakeven_month": r["breakeven_month"],
                "ending_cash": r["ending_cash"],
            })
        rows.append({"driver": key, "label": label, "cells": cells})
    return {"deltas_pct": [int(d * 100) for d in deltas], "rows": rows}


def _capital_recompute(a: Assumptions, computed: dict) -> dict:
    """Map computed metrics onto the v2 scoring engine's capital category.

    The engine (`services/scoring.py::SCORING_V2_WEIGHTS["capital"]`) wants
    two 0..10 sliders: `runway` and `burn_efficiency`. We derive both from
    the financial model so the founder can see how their assumptions move
    the capital score before the next scoring run picks them up.
    """
    runway = float(computed.get("runway_months") or 0)
    runway_slider = max(0.0, min(10.0, runway / 2.4))  # 24mo+ = 10/10

    ratio = computed.get("ltv_cac_ratio")
    if ratio is None:
        burn_slider = 5.0
    else:
        burn_slider = max(0.0, min(10.0, float(ratio) * 3.0))  # 3.3:1 = 10/10

    burn_pts = round((burn_slider / 10.0) * 5, 2)
    runway_pts = round((runway_slider / 10.0) * 5, 2)
    total = round(min(burn_pts + runway_pts, 10), 2)

    return {
        "category": "capital",
        "max": 10,
        "total": total,
        "factors": {
            "burn_efficiency": {
                "raw": round(burn_slider, 2),
                "points": burn_pts,
                "max": 5,
                "label": "Burn efficiency",
                "source": "ltv_cac_ratio",
            },
            "runway": {
                "raw": round(runway_slider, 2),
                "points": runway_pts,
                "max": 5,
                "label": "Runway / unit econ.",
                "source": "runway_months",
            },
        },
    }


# ---------------------------------------------------------------------------
# Authorization helpers — mirrors the pattern used by scoring/legal routes.
#   • admin / partner / investor → privileged: full read; admin can also edit.
#   • founder → may read & edit only the project they own (`founder_id` match).
#   • everyone else → blocked.
# ---------------------------------------------------------------------------
def _ensure_can_view(project: Project, user: User) -> None:
    if is_privileged(user):
        return
    if not can_access_founder_resource(user, project.founder_id):
        raise HTTPException(status_code=403, detail="Forbidden: you do not own this project")


def _ensure_can_edit(project: Project, user: User) -> None:
    if user.role == UserRole.ADMIN:
        return
    if user.role == UserRole.FOUNDER:
        if not can_access_founder_resource(user, project.founder_id):
            raise HTTPException(status_code=403, detail="Forbidden: you do not own this project")
        return
    raise HTTPException(status_code=403, detail="Read-only for non-founder roles")


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@router.get("/{project_id}")
def get_model(
    project_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    project = session.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_can_view(project, user)

    fm = session.exec(
        select(FinancialModel).where(FinancialModel.project_id == project_id)
    ).first()

    if not fm:
        # Return a default scaffold without persisting yet.
        a = Assumptions()
        computed = _project_financials(a)
        return {
            "exists": False,
            "project_id": project_id,
            "project_name": project.name,
            "assumptions": a.model_dump(),
            "computed": computed,
            "sensitivity": _sensitivity(a),
            "capital_recompute": _capital_recompute(a, computed),
            "updated_at": None,
        }

    return {
        "exists": True,
        "project_id": project_id,
        "project_name": project.name,
        "assumptions": json.loads(fm.assumptions_json or "{}"),
        "computed": json.loads(fm.computed_json or "{}"),
        "sensitivity": json.loads(fm.sensitivity_json or "{}"),
        "capital_recompute": json.loads(fm.capital_recompute_json or "{}") if fm.capital_recompute_json else None,
        "updated_at": fm.updated_at.isoformat() if fm.updated_at else None,
    }


class UpsertRequest(BaseModel):
    assumptions: Assumptions


@router.put("/{project_id}")
def upsert_model(
    project_id: int,
    body: UpsertRequest,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    project = session.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_can_edit(project, user)

    a = body.assumptions
    computed = _project_financials(a)
    sensitivity = _sensitivity(a)
    capital = _capital_recompute(a, computed)

    fm = session.exec(
        select(FinancialModel).where(FinancialModel.project_id == project_id)
    ).first()
    if not fm:
        fm = FinancialModel(project_id=project_id)
        session.add(fm)

    fm.assumptions_json = json.dumps(a.model_dump())
    fm.computed_json = json.dumps(computed)
    fm.sensitivity_json = json.dumps(sensitivity)
    fm.capital_recompute_json = json.dumps(capital)
    fm.updated_by = user.id
    fm.updated_at = datetime.utcnow()
    session.add(fm)

    session.add(ActivityLog(
        action="financial_model_saved",
        details=(
            f"Project {project.name}: runway={computed['runway_months']}mo, "
            f"breakeven={computed['breakeven_month']}, capital_score={capital['total']}/10"
        ),
        actor=user.email,
    ))
    session.commit()

    return {
        "exists": True,
        "project_id": project_id,
        "project_name": project.name,
        "assumptions": a.model_dump(),
        "computed": computed,
        "sensitivity": sensitivity,
        "capital_recompute": capital,
        "updated_at": fm.updated_at.isoformat(),
    }


@router.post("/{project_id}/recompute")
def recompute_model(
    project_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    project = session.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_can_view(project, user)

    fm = session.exec(
        select(FinancialModel).where(FinancialModel.project_id == project_id)
    ).first()
    if not fm:
        raise HTTPException(status_code=404, detail="No financial model saved yet")

    a = Assumptions(**json.loads(fm.assumptions_json or "{}"))
    computed = _project_financials(a)
    sensitivity = _sensitivity(a)
    capital = _capital_recompute(a, computed)

    fm.computed_json = json.dumps(computed)
    fm.sensitivity_json = json.dumps(sensitivity)
    fm.capital_recompute_json = json.dumps(capital)
    fm.updated_at = datetime.utcnow()
    session.add(fm)
    session.commit()

    return {
        "computed": computed,
        "sensitivity": sensitivity,
        "capital_recompute": capital,
    }


@router.get("/{project_id}/export.xlsx")
def export_xlsx(
    project_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    project = session.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    _ensure_can_view(project, user)

    fm = session.exec(
        select(FinancialModel).where(FinancialModel.project_id == project_id)
    ).first()
    if fm:
        a = Assumptions(**json.loads(fm.assumptions_json or "{}"))
    else:
        a = Assumptions()
    computed = _project_financials(a)
    sensitivity = _sensitivity(a)
    capital = _capital_recompute(a, computed)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"openpyxl not available: {exc}")

    wb = Workbook()

    # --- Assumptions sheet
    ws = wb.active
    ws.title = "Assumptions"
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEE9FE")
    ws.append(["Driver", "Value"])
    for c in ws[1]:
        c.font = bold
        c.fill = header_fill
    for k, v in a.model_dump().items():
        ws.append([k, v])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # --- Projection sheet
    ws2 = wb.create_sheet("Projection")
    headers = ["Month", "Units", "Revenue", "Gross Profit", "Marketing", "Fixed Cost", "Net", "Cash"]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = bold
        c.fill = header_fill
    for m in computed["months"]:
        ws2.append([
            m["month"], m["units"], m["revenue"], m["gross_profit"],
            m["marketing"], m["fixed_cost"], m["net"], m["cash"],
        ])
    for col in "ABCDEFGH":
        ws2.column_dimensions[col].width = 14

    # --- Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    for c in ws3[1]:
        c.font = bold
        c.fill = header_fill
    summary_rows = [
        ("Runway (months)", computed["runway_months"]),
        ("Avg monthly burn", computed["avg_monthly_burn"]),
        ("Breakeven month", computed["breakeven_month"] or "Not within horizon"),
        ("Ending cash", computed["ending_cash"]),
        ("Total revenue (horizon)", computed["total_revenue_horizon"]),
        ("LTV", computed["ltv"]),
        ("LTV / CAC", computed["ltv_cac_ratio"]),
        ("Capital score (recompute)", f"{capital['total']} / 10"),
    ]
    for k, v in summary_rows:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 22

    # --- Sensitivity sheet
    ws4 = wb.create_sheet("Sensitivity")
    ws4.append(["Driver"] + [f"{d:+d}% (runway mo)" for d in sensitivity["deltas_pct"]])
    for c in ws4[1]:
        c.font = bold
        c.fill = header_fill
    for row in sensitivity["rows"]:
        ws4.append([row["label"]] + [c["runway_months"] for c in row["cells"]])
    ws4.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws4.column_dimensions[col].width = 18

    # Title row above each sheet
    for sheet in (ws, ws2, ws3, ws4):
        sheet.insert_rows(1)
        sheet["A1"] = f"{project.name} — Financial Model"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="left")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in project.name)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_financials.xlsx"'},
    )
